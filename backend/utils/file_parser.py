import csv
import io
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from backend.core.exceptions import app_validation_error


def parse_excel_to_list(file_content: bytes) -> list[dict[str, Any]]:
    """解析 Excel 文件为 list[dict]"""
    try:
        workbook = load_workbook(
            filename=BytesIO(file_content), read_only=True, data_only=True
        )
        sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).strip() for h in rows[0]]
        data: list[dict[str, Any]] = []

        for row in rows[1:]:
            if not any(row):
                continue

            row_dict = dict(zip(headers, row, strict=False))
            data.append(row_dict)

        return data
    except Exception as exc:
        raise app_validation_error(
            f"Excel 解析失败: {exc}",
            code="FILE_PARSE_ERROR",
        ) from exc


def parse_csv_to_list(file_content: bytes) -> list[dict[str, Any]]:
    """解析 CSV 文件为 list[dict]"""
    try:
        content_str = file_content.decode("utf-8-sig")
        f = io.StringIO(content_str)
        reader = csv.DictReader(f)
        return [row for row in reader]
    except UnicodeDecodeError as exc:
        raise app_validation_error(
            "文件编码格式不正确，请上传 UTF-8 编码的文件",
            code="FILE_ENCODING_ERROR",
        ) from exc
    except csv.Error as exc:
        raise app_validation_error(f"CSV 解析失败: {exc}", code="FILE_PARSE_ERROR") from exc
    except Exception as exc:
        raise app_validation_error(f"CSV 解析失败: {exc}", code="FILE_PARSE_ERROR") from exc


def parse_file(filename: str, file_content: bytes) -> list[dict[str, Any]]:
    """总入口：根据文件名后缀选择解析器"""
    suffix = Path(filename).suffix.lower()
    if suffix == ".xlsx":
        return parse_excel_to_list(file_content)
    if suffix == ".xls":
        raise app_validation_error(
            "暂不支持 .xls，请转换为 .xlsx 后再上传",
            code="UNSUPPORTED_FILE_TYPE",
        )
    if suffix == ".csv":
        return parse_csv_to_list(file_content)
    raise app_validation_error(
        "不支持的文件格式，仅支持 .xlsx, .csv",
        code="UNSUPPORTED_FILE_TYPE",
    )
