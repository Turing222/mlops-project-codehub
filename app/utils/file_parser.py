# utils/file_parser.py
import csv
import functools
import io
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.core.exceptions import FileParseException


# 定义一个异常，方便上层捕获处理
# 1. 定义：装饰器本身是一个函数，但其内部的 wrapper 必须是 async
def handle_file_exceptions(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        try:
            # 执行被装饰的异步函数
            return func(*args, **kwargs)
        except UnicodeDecodeError as e:
            raise FileParseException(
                "文件编码格式不正确，请上传 UTF-8 编码的文件"
            ) from e
        except ValueError as e:
            raise FileParseException(f"文件内容校验失败: {str(e)}") from e
        except Exception as e:
            raise FileParseException("发生未知文件导入异常") from e

    return wrapper


@handle_file_exceptions
def parse_excel_to_list(file_content: bytes) -> list[dict[str, Any]]:
    """解析 Excel 文件为 list[dict]"""
    try:
        # 使用 BytesIO 将二进制数据转换为内存文件对象
        workbook = load_workbook(
            filename=BytesIO(file_content), read_only=True, data_only=True
        )
        sheet = workbook.active

        # 假设第一行是表头 (header)
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [
            str(h).strip() for h in rows[0]
        ]  # 获取表头，比如 ['username', 'email']
        data = []

        # 从第二行开始遍历数据
        for row in rows[1:]:
            # 简单的行数据清洗：过滤全空行
            if not any(row):
                continue

            # 将每一行映射为字典: {'username': 'xxx', 'email': 'xxx'}
            # zip(headers, row) 会自动把表头和值对应起来
            row_dict = dict(zip(headers, row, strict=False))
            data.append(row_dict)

        return data
    except Exception as e:
        raise FileParseException(f"Excel 解析失败: {str(e)}") from e


@handle_file_exceptions
def parse_csv_to_list(file_content: bytes) -> list[dict[str, Any]]:
    """解析 CSV 文件为 list[dict]"""
    try:
        # 将 bytes 解码为 string，CSV 模块处理的是文本
        content_str = file_content.decode("utf-8-sig")  # utf-8-sig 可以自动处理 BOM 头
        f = io.StringIO(content_str)

        reader = csv.DictReader(f)
        data = [row for row in reader]
        return data
    except Exception as e:
        raise FileParseException(f"CSV 解析失败: {str(e)}") from e


def parse_file(filename: str, file_content: bytes) -> list[dict[str, Any]]:
    """总入口：根据文件名后缀选择解析器"""
    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        return parse_excel_to_list(file_content)
    elif filename.endswith(".csv"):
        return parse_csv_to_list(file_content)
    else:
        raise FileParseException("不支持的文件格式，仅支持 .xlsx, .xls, .csv")
